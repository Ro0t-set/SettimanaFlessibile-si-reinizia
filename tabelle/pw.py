import string
from random import *
import openpyxl
from django.contrib.auth.models import User
characters =  string.digits
riga = 0
while riga < 1417:
        riga = riga + 1
        password =  "".join(choice(characters) for x in range(randint(8, 8)))
        excel_document = openpyxl.load_workbook('listautenti.xlsx')
        sheet = excel_document.get_sheet_by_name('Foglio1')
        nome_utente= str(sheet.cell(row = riga, column = 1).value)
        classe= str(sheet.cell(row = riga, column = 2).value)
        utente=str(nome_utente+classe)
        print (riga)
        print (utente)
        print (password)
        user=User.objects.create_user(utente, password=password)
        user.is_superuser=False
        user.is_staff=False
        user.save()

      


import string
from random import *
import openpyxl
from django.contrib.auth.models import User
from app.models import Iscrizione
characters =  string.digits
riga = 0
while riga < 1417:
        riga = riga + 1
        excel_document = openpyxl.load_workbook('listautenti.xlsx')
        sheet = excel_document.get_sheet_by_name('Foglio1')
        nome_utente= (sheet.cell(row = riga, column = 1).value)
        classe= (sheet.cell(row = riga, column = 2).value)
        utente=(nome_utente+classe)
        print (riga)
        utente = User.objects.get(username=utente)
        Iscrizione.objects.create(user=utente)